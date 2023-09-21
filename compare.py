import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill
# import sys
# import git
# from git import Repo

def compare_excel_files(x_file, y_file, result_file):

    file1 = pd.ExcelFile(x_file)
    file2 = pd.ExcelFile(y_file)

    sheets1 = file1.sheet_names
    sheets2 = file2.sheet_names

    common_sheets = set(sheets1) & set(sheets2)

    workbook = Workbook()

    for sheet_name in common_sheets:

        df1 = pd.read_excel(file1, sheet_name)
        df2 = pd.read_excel(file2, sheet_name)

        # Defining the unique value for the comparison of rows
        primaryKey = df1.columns[0]
        added_rows = df2[~df2[primaryKey].isin(df1[primaryKey])]
        deleted_rows = df1[~df1[primaryKey].isin(df2[primaryKey])]
        modified_rows = []

        for _, row2 in df2.iterrows():
            pk = row2[primaryKey]
            row1 = df1[df1[primaryKey] == pk]
            if row1.empty:
                continue
            elif not (row1.iloc[0][1:].equals(row2[1:])):
                modified_rows.append(row2.tolist())
                modified_rows.append(row1.iloc[0].tolist())
                modified_rows.append('')

        header = df1.columns.tolist()
        result_df = pd.DataFrame(columns=header)
        # Header
        result_df.loc[0] = header

        # Add recently added rows
        if not added_rows.empty:
            result_df = result_df._append(pd.Series(['Added'] + [''] * (len(header) - 1), index=header),
                                          ignore_index=True)
            result_df = result_df._append(added_rows, ignore_index=True)
            result_df = result_df._append(pd.Series([''] * len(header), index=header),
                                          ignore_index=True)  # Empty row

        # Add recently deleted rows
        if not deleted_rows.empty:
            result_df = result_df._append(pd.Series(['Deleted'] + [''] * (len(header) - 1), index=header),
                                          ignore_index=True)
            result_df = result_df._append(deleted_rows, ignore_index=True)
            result_df = result_df._append(pd.Series([''] * len(header), index=header),
                                          ignore_index=True)  # Empty row
 
        # Add recently modified rows
        if modified_rows:
            result_df = result_df._append(pd.Series(['Modified'] + [''] * (len(header) - 1), index=header),
                                          ignore_index=True)
            result_df = result_df._append(pd.DataFrame(modified_rows, columns=header), ignore_index=True)
 
        modified_rows_df = pd.DataFrame(modified_rows)
        worksheet = workbook.create_sheet(sheet_name)

        for row in dataframe_to_rows(result_df, index=False, header=False):
            worksheet.append(row)

 

        # Uages: iteration of the row for highlighting the modified cell
        row_index = 0
        end = len(modified_rows)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):

            if row[0].value in added_rows[primaryKey].tolist():
                for cell in row:
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00',
                                            fill_type='solid')  # Green

            elif row[0].value in deleted_rows[primaryKey].tolist():
                for cell in row:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                            fill_type='solid')  # Red
            elif row[0].value in modified_rows_df[0].tolist():
                cell_index = 0
                for cell in row:
                    if ((row_index < end - 1 and cell_index < 8) and (row_index == 0 or row_index % 3 == 0) and (
                            cell.value != modified_rows[row_index + 1][cell_index]) and not pd.isna(cell.value)):
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                                                fill_type='solid')  # Yellow
                    cell_index = cell_index + 1
                row_index = row_index + 1

    workbook.remove(workbook['Sheet'])
    workbook.save(result_file)

# # File1 should be the older version and File2 should be the latest version of the file.

file1 = './new_version/DTE_ARTEMIS.xlsx'
file2 = './old_version/DTE_ARTEMIS.xlsx'
resultFile = 'REPORT.xlsx'

compare_excel_files(file1, file2, resultFile)


# # Folders containing old and new versions of Excel files
# old_folder = 'old_version'
# new_folder = 'new_version'
# result_folder = 'comparison_results'

# # Ensure the result folder exists
# if not os.path.exists(result_folder):
#     os.mkdir(result_folder)

# # Loop through files in both folders and compare them
# # for root, dirs, files in os.walk(old_folder):
# #     for file in files:
# #         if file.endswith(".xlsx"):
# #             old_file_path = os.path.join(root, file)
# #             new_file_path = os.path.join(new_folder, file)
# #             result_file_path = os.path.join(result_folder, file)
# #             compare_excel_files(old_file_path, new_file_path, result_file_path)


# if len(sys.argv) != 3:
#     print("Usage: python script.py <branch1> <branch2>")
#     sys.exit(1)

# branch1 = sys.argv[1]
# branch2 = sys.argv[2]

# repo_path = "https://git.i.mercedes-benz.com/SHAHTEJ/excel_comparision"

# # repo = git.Repo(repo_path)
# repo=Repo.clone_from(repo_path, 'DTE')
# repo.git.checkout(branch1)
# dte_folder_path_branch1 = os.path.join(repo_path, "DTE")


# repo.git.checkout(branch2)
# dte_folder_path_branch2 = os.path.join(repo_path, "DTE")

# excel_files_branch1 = [f for f in os.listdir(dte_folder_path_branch1) if f.endswith(".xlsx")]
# excel_files_branch2 = [f for f in os.listdir(dte_folder_path_branch2) if f.endswith(".xlsx")]

# file_details = []

# for file1, file2 in zip(excel_files_branch1, excel_files_branch2):
#     file_path1 = os.path.join(dte_folder_path_branch1, file1)
#     file_path2 = os.path.join(dte_folder_path_branch2, file2)
    
#     comparison_result = compare_excel_files(file_path1, file_path2)